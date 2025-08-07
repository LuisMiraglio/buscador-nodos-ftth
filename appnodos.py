import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import openpyxl
import os
import re

class NodoApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Buscador de Nodos FTTH")
        self.root.geometry("900x700")
        self.root.configure(bg="#0f172a")  # Fondo azul marino profundo
        self.root.minsize(800, 600)
        
        # Cargar el icono de la aplicación
        icon_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "lupa.ico")
        try:
            self.root.iconbitmap(icon_path)
        except Exception as e:
            print(f"No se pudo cargar el icono: {str(e)}")
        
        # Fuentes personalizadas mejoradas
        title_font = ("Segoe UI", 20, "bold")
        subtitle_font = ("Segoe UI", 11, "bold")
        label_font = ("Segoe UI", 10)
        result_font = ("JetBrains Mono", 10)  # Fuente monoespaciada más moderna
        
        self.excel_data = None
        self.excel_path = None
        
        # Ruta predeterminada para los archivos Excel
        self.excel_folder = r"C:\Users\LMIRAGLIO\Excel-luis"
        
        # Crear estilo personalizado con diseño moderno
        style = ttk.Style()
        style.theme_use('clam')
        
        # Paleta de colores moderna
        colors = {
            'primary': '#3b82f6',      # Azul vibrante
            'primary_hover': '#2563eb', # Azul más oscuro
            'success': '#10b981',      # Verde moderno
            'success_hover': '#059669', # Verde más oscuro
            'danger': '#ef4444',       # Rojo moderno
            'warning': '#f59e0b',      # Naranja/amarillo
            'dark': '#1e293b',         # Gris oscuro
            'light': '#f8fafc',        # Blanco suave
            'accent': '#8b5cf6'        # Púrpura
        }
        
        # Botones con diseño moderno y sombras
        style.configure("Modern.TButton", 
                  font=("Segoe UI", 10, "bold"),
                  background=colors['primary'],
                  foreground="white",
                  borderwidth=0,
                  focuscolor='none',
                  relief="flat",
                  padding=(20, 10))
        style.map("Modern.TButton",
                  background=[("active", colors['primary_hover']),
                            ("pressed", colors['primary_hover'])])
    
        # Botón de búsqueda principal
        style.configure("Search.TButton", 
                  font=("Segoe UI", 11, "bold"),
                  background=colors['primary'],
                  foreground="white",
                  borderwidth=0,
                  focuscolor='none',
                  relief="flat",
                  padding=(25, 12))
        style.map("Search.TButton",
                  background=[("active", colors['primary_hover']),
                            ("pressed", colors['primary_hover'])])
    
        # Botón de carga con color distintivo
        style.configure("Load.TButton", 
                  font=("Segoe UI", 10, "bold"),
                  background=colors['success'],
                  foreground="white",
                  borderwidth=0,
                  focuscolor='none',
                  relief="flat",
                  padding=(20, 10))
        style.map("Load.TButton",
                  background=[("active", colors['success_hover']),
                            ("pressed", colors['success_hover'])])
    
        # Campo de entrada con diseño moderno
        style.configure("Modern.TEntry", 
                  font=("Segoe UI", 11),
                  fieldbackground="#ffffff",
                  borderwidth=2,
                  relief="flat",
                  focuscolor=colors['primary'],
                  insertcolor="#000000")
        
        
        
        # Frame principal con padding mejorado
        self.main_frame = tk.Frame(self.root, padx=40, pady=30, bg="#0f172a")
        self.main_frame.pack(fill=tk.BOTH, expand=True)
        
        # Header con gradiente visual y diseño moderno
        header_frame = tk.Frame(self.main_frame, bg="#1e40af", pady=20, relief=tk.FLAT)
        header_frame.pack(fill=tk.X, pady=(0, 25))
        
        # Container para el título y subtitle
        title_container = tk.Frame(header_frame, bg="#1e40af")
        title_container.pack(expand=True)
        
        self.title_label = tk.Label(title_container, text="🔍 Buscador de Nodos FTTH", 
                                  font=title_font, bg="#1e40af", fg="#ffffff")
        self.title_label.pack(pady=(0, 5))
        

        
        # Frame para archivo con diseño de tarjeta
        self.file_frame = tk.Frame(self.main_frame, bg="#1e293b", pady=15, relief=tk.FLAT, bd=2)
        self.file_frame.pack(fill=tk.X, pady=(0, 20))
        
        # Contenedor interno para mejor organización
        file_inner_frame = tk.Frame(self.file_frame, bg="#1e293b")
        file_inner_frame.pack(fill=tk.X, padx=20, pady=10)
        
        # Etiqueta de sección
        file_section_label = tk.Label(file_inner_frame, text="📁 Archivo Excel", 
                                    font=subtitle_font, bg="#1e293b", fg="#f1f5f9")
        file_section_label.pack(side=tk.LEFT, padx=(0, 15))
        
        # Botón para cargar archivo Excel con estilo moderno
        self.load_button = ttk.Button(
            file_inner_frame, 
            text="Cargar Archivo", 
            command=self.load_excel,
            style="Load.TButton"
        )
        self.load_button.pack(side=tk.LEFT, padx=(0, 15))
        
        # Etiqueta del archivo cargado con mejor diseño
        self.file_label = tk.Label(
            file_inner_frame, 
            text="Ningún archivo cargado", 
            font=label_font,
            fg="#94a3b8",
            bg="#1e293b",
            padx=10
        )
        self.file_label.pack(side=tk.LEFT, fill=tk.X, expand=True)
        
        # Frame de búsqueda con diseño de tarjeta moderna
        self.search_frame = tk.Frame(self.main_frame, bg="#1e293b", pady=20, relief=tk.FLAT, bd=2)
        self.search_frame.pack(fill=tk.X, pady=(0, 20))
        
        # Contenedor interno para la búsqueda
        search_inner_frame = tk.Frame(self.search_frame, bg="#1e293b")
        search_inner_frame.pack(padx=20, pady=10)
        
        # Etiqueta de sección de búsqueda
        search_section_label = tk.Label(search_inner_frame, text="🔍 Búsqueda de Nodo", 
                                      font=subtitle_font, bg="#1e293b", fg="#f1f5f9")
        search_section_label.pack(pady=(0, 15))
        
        # Container para entrada y botón
        search_controls = tk.Frame(search_inner_frame, bg="#1e293b")
        search_controls.pack()
        
        self.search_label = tk.Label(
            search_controls, 
            text="Número de Nodo:", 
            font=("Segoe UI", 11, "bold"),
            bg="#1e293b",
            fg="#e2e8f0"
        )
        self.search_label.pack(side=tk.LEFT, padx=(0, 15))
        
        # Campo de entrada con diseño moderno
        entry_frame = tk.Frame(search_controls, bg="#ffffff", relief=tk.FLAT, bd=2)
        entry_frame.pack(side=tk.LEFT, padx=(0, 15))
        
        self.search_entry = tk.Entry(
            entry_frame, 
            width=20, 
            font=("Segoe UI", 12),
            bg="#ffffff",
            fg="#1e293b",
            relief=tk.FLAT,
            bd=0,
            insertbackground="#3b82f6",
            highlightthickness=0
        )
        self.search_entry.pack(padx=10, pady=8)
        
        # Botón de búsqueda con estilo moderno
        self.search_button = ttk.Button(
            search_controls, 
            text="🔍 Buscar", 
            command=self.search_node,
            style="Search.TButton"
        )
        self.search_button.pack(side=tk.LEFT)
        
        # Búsqueda rápida con Enter
        self.search_entry.bind('<Return>', lambda event: self.search_node())
        
        # Frame de resultados optimizado para mejor visualización
        self.results_frame = tk.LabelFrame(
            self.main_frame, 
            text=" 📊 Resultados de Búsqueda ", 
            padx=15, 
            pady=10,
            bg="#1e293b",
            font=("Segoe UI", 12, "bold"),
            fg="#f1f5f9",
            relief=tk.FLAT,
            bd=2,
            labelanchor='n'
        )
        self.results_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 10))
        
        # Crear un marco para el texto y scrollbar más compacto
        text_frame = tk.Frame(self.results_frame, bg="#1e293b")
        text_frame.pack(fill=tk.BOTH, expand=True, padx=8, pady=5)
        
        # Área de resultados optimizada para mostrar todo el contenido
        self.results_text = tk.Text(
            text_frame, 
            height=8, 
            width=80, 
            font=result_font,
            wrap=tk.WORD,
            borderwidth=0,
            relief=tk.FLAT,
            bg="#f8fafc",
            fg="#1e293b",
            insertbackground="#3b82f6",
            padx=15,
            pady=10,
            selectbackground="#bfdbfe",
            selectforeground="#1e40af",
            highlightthickness=0,
            spacing1=2,
            spacing2=1,
            spacing3=2
        )
        self.results_text.config(state="disabled")
        self.results_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
         
        # Status bar moderno con gradiente
        status_frame = tk.Frame(self.root, bg="#1e40af", height=35)
        status_frame.pack(side=tk.BOTTOM, fill=tk.X)
        status_frame.pack_propagate(False)
        
        self.status = tk.Label(
            status_frame, 
            bd=0, 
            relief=tk.FLAT, 
            anchor=tk.W,
            font=("Segoe UI", 10, "bold"),
            bg="#1e40af",
            fg="#ffffff",
            padx=20,
            pady=8
        )
        self.status.pack(fill=tk.BOTH, expand=True)
        
    def load_excel(self):
        self.excel_path = filedialog.askopenfilename(
            title="Seleccionar archivo Excel",
            initialdir=self.excel_folder,  # Usar la carpeta definida para Excel
            filetypes=(("Excel files", "*.xlsx *.xls"), ("All files", "*.*"))
        )
        
        if not self.excel_path:
            return
        
        try:
            # Configurar etiquetas para formateo de texto compacto y legible
            self.results_text.tag_configure("success", 
                                           foreground="#059669", 
                                           font=("Segoe UI", 12, "bold"),
                                           spacing1=3, spacing3=2)
            self.results_text.tag_configure("info", 
                                           foreground="#0369a1", 
                                           font=("Segoe UI", 10))
            self.results_text.tag_configure("info_value", 
                                           foreground="#1e40af", 
                                           font=("JetBrains Mono", 10, "bold"),
                                           background="#eff6ff")
            self.results_text.tag_configure("info_highlight", 
                                           foreground="#7c3aed", 
                                           font=("Segoe UI", 10, "bold"),
                                           background="#f3e8ff")
            self.results_text.tag_configure("warning", 
                                           foreground="#d97706", 
                                           font=("Segoe UI", 10, "bold"))
            self.results_text.tag_configure("error", 
                                           foreground="#dc2626", 
                                           font=("Segoe UI", 10, "bold"))
            self.results_text.tag_configure("label", 
                                           foreground="#374151", 
                                           font=("Segoe UI", 10, "bold"))
            self.results_text.tag_configure("separator", 
                                           foreground="#6b7280", 
                                           font=("Segoe UI", 9),
                                           spacing1=2, spacing3=2)
            
            self.status.config(text="⏳ Cargando archivo...")
            self.root.update()
            
            # Cargar el archivo Excel con openpyxl
            self.excel_data = openpyxl.load_workbook(self.excel_path, data_only=True)
            filename = os.path.basename(self.excel_path)
            self.file_label.config(text=f"✅ {filename}", fg="#10b981")
            self.status.config(text=f"✅ Archivo {filename} cargado correctamente")
            
        except Exception as e:
            self.status.config(text="❌ Error al cargar archivo")
            messagebox.showerror("Error", f"Error al cargar el archivo: {str(e)}")
    
    def search_node(self):
        if not self.excel_data:
            messagebox.showwarning("Advertencia", "Primero debes cargar un archivo Excel")
            return
        
        node_to_search = self.search_entry.get().strip()
        if not node_to_search:
            messagebox.showwarning("Advertencia", "Ingresa un número de nodo para buscar")
            return
        
        # Limpiar resultados anteriores
        self.results_text.config(state="normal")
        self.results_text.delete(1.0, tk.END)
        self.status.config(text="🔍 Buscando nodo " + node_to_search + "...")
        self.root.update()
        
        results_found = False
        
        try:
            # Primero intentar en la hoja "Caja Central" que tiene la estructura principal
            if "Caja Central" in self.excel_data.sheetnames:
                sheet = self.excel_data["Caja Central"]
                
                # Buscar en la estructura específica de esta hoja
                results_found = self.search_in_caja_central(sheet, node_to_search)
            
            # Si no se encuentra en Caja Central, buscar en otras hojas
            if not results_found:
                for sheet_name in self.excel_data.sheetnames:
                    if sheet_name != "Caja Central":
                        sheet = self.excel_data[sheet_name]
                        found = self.search_in_other_sheet(sheet, sheet_name, node_to_search)
                        if found:
                            results_found = True
            
            if not results_found:
                self.results_text.insert(tk.END, "🔍 ", "info")
                self.results_text.insert(tk.END, "Resultado de búsqueda\n", "label")
                self.results_text.insert(tk.END, "─" * 60 + "\n", "separator")
                self.results_text.insert(tk.END, "❌ No se encontró el nodo '", "error")
                self.results_text.insert(tk.END, node_to_search, "info_value")
                self.results_text.insert(tk.END, "' en el archivo\n", "error")
                self.results_text.insert(tk.END, "\n💡 Sugerencias:\n", "label")
                self.results_text.insert(tk.END, "• Verifica que el número de nodo sea correcto\n", "info")
                self.results_text.insert(tk.END, "• Asegúrate de que el archivo Excel contiene los datos\n", "info")
                self.results_text.insert(tk.END, "• Intenta buscar sin espacios adicionales", "info")
                
            self.status.config(text="✅ Búsqueda completada para nodo " + node_to_search)
            self.results_text.config(state="disabled")
                
        except Exception as e:
            self.status.config(text=f"Error durante la búsqueda: {str(e)}")
            messagebox.showerror("Error", f"Error al buscar: {str(e)}")
        
        self.results_text.config(state="disabled")  # Volver a hacer el área de resultados de solo lectura

    def get_olt_info(self, cell, color_text=None):
        """Determina a qué OLT pertenece un nodo basado en el color de fondo de la celda"""
        # Primero intentar detectar por el color de fondo de la celda
        try:
            if cell and hasattr(cell, 'fill') and cell.fill:
                fill = cell.fill
                if hasattr(fill, 'fgColor') and fill.fgColor:
                    # Verde puede estar codificado de varias maneras
                    if hasattr(fill.fgColor, 'rgb') and fill.fgColor.rgb:
                        rgb_value = fill.fgColor.rgb
                        print(f"RGB value: {rgb_value}")
                        # Valores comunes para verde
                        if rgb_value and ('00FF00' in rgb_value or 'FF00FF00' in rgb_value or '92D050' in rgb_value):
                            return "OLT2"
                    
                    # Verificar también valores de tema y tint que puedan indicar verde
                    if hasattr(fill.fgColor, 'theme') and fill.fgColor.theme:
                        theme = fill.fgColor.theme
                        tint = getattr(fill.fgColor, 'tint', 0)
                        print(f"Theme: {theme}, Tint: {tint}")
                        # Tema 6 suele ser verde
                        if theme == 6:
                            return "OLT2"
        except Exception as e:
            print(f"Error al obtener color de celda: {e}")
            
        # Si no podemos determinar por el fondo, intentar por el texto
        if color_text:
            color_lower = str(color_text).lower().strip()
            if 'verde' in color_lower:
                return "OLT2"
        
        # Por defecto es OLT1
        return "OLT1"

    def search_in_caja_central(self, sheet, node_to_search):
        """Busca un nodo en la hoja Caja Central con su estructura específica"""
        results_found = False
        
        # En esta hoja, los encabezados están en la fila 2, los colores en la fila 3,
        # y las ubicaciones en la fila 1
        
        # Buscar todas las columnas que contienen "Nodo"
        nodo_columns = []
        for col in range(1, sheet.max_column + 1):
            if str(sheet.cell(row=2, column=col).value).lower().strip() == "nodo":
                nodo_columns.append(col)
        
        # Buscar el nodo específico en las filas de datos
        for row in range(4, sheet.max_row + 1):
            for nodo_col in nodo_columns:
                cell = sheet.cell(row=row, column=nodo_col)
                cell_value = str(cell.value or "").strip()
                
                # Si encontramos una coincidencia exacta
                if cell_value == node_to_search:
                    results_found = True
                    
                    # Obtener la placa y puerto
                    placa_value = sheet.cell(row=row, column=nodo_col + 1).value
                    port_value = sheet.cell(row=row, column=nodo_col + 2).value
                    
                    # Obtener el color de fila (columna 1)
                    row_color = sheet.cell(row=row, column=1).value
                    color_str = str(row_color or "").strip()
                    
                    # Determinar a qué OLT pertenece basado en el color DE FONDO de la celda del nodo
                    olt_info = self.get_olt_info(cell, color_str)
                    
                    # Mostrar resultados de forma compacta y clara
                    self.results_text.insert(tk.END, "🎯 ", "success")
                    self.results_text.insert(tk.END, "Nodo encontrado: ", "success")
                    self.results_text.insert(tk.END, f"{node_to_search}\n", "info_value")
                    self.results_text.insert(tk.END, "─" * 60 + "\n", "separator")
                    
                    # Mostrar información en formato compacto de tabla
                    self.results_text.insert(tk.END, "🔧 Placa: ", "label")
                    self.results_text.insert(tk.END, f"{placa_value}   ", "info_value")
                    self.results_text.insert(tk.END, "🔌 Puerto: ", "label")
                    self.results_text.insert(tk.END, f"{port_value}   ", "info_value")
                    self.results_text.insert(tk.END, "🌐 OLT: ", "label")
                    self.results_text.insert(tk.END, f"{olt_info}", "info_highlight")
        
        return results_found

    # Agregar una función específica para determinar color por celda
    def get_cell_color(self, sheet, row, col):
        """Intenta determinar el color de una celda de varias formas posibles"""
        # 1. Primero intentar obtener el valor directamente
        color = str(sheet.cell(row=row, column=col).value or "").strip()
        
        # 2. Intentar obtener el color del estilo de la celda si existe
        try:
            cell = sheet.cell(row=row, column=col)
            if hasattr(cell, 'fill') and cell.fill:
                fill = cell.fill
                if hasattr(fill, 'fgColor') and fill.fgColor:
                    # Verificar si hay un color de relleno
                    if hasattr(fill.fgColor, 'rgb') and fill.fgColor.rgb:
                        rgb = fill.fgColor.rgb
                        # Verde suele contener mucho componente G (green)
                        if rgb and 'FF00FF00' in rgb:  # Verde brillante en hex
                            return "Verde"
                    elif hasattr(fill.fgColor, 'theme') and fill.fgColor.theme:
                        # Los colores de tema 6 suelen ser verdes en Excel
                        if fill.fgColor.theme == 6:
                            return "Verde"
        except:
            # Si hay error al obtener el color, seguimos con el valor de texto
            pass
            
        return color
    
    def search_in_other_sheet(self, sheet, sheet_name, node_to_search):
        """Busca un nodo en otras hojas del Excel con estructura variable"""
        results_found = False
        
        # Implementación genérica para otras hojas
        # Buscamos primero columnas que contengan "Nodo" en cualquier fila
        nodo_col_candidates = []
        for row in range(1, min(15, sheet.max_row + 1)):
            for col in range(1, min(30, sheet.max_column + 1)):
                cell_value = str(sheet.cell(row=row, column=col).value or "").strip().lower()
                if "nodo" in cell_value:
                    nodo_col_candidates.append((row, col))
        
        # Para cada candidato, buscar el nodo en las filas siguientes
        for header_row, nodo_col in nodo_col_candidates:
            # Buscar en todas las filas por debajo del encabezado
            for row in range(header_row + 1, sheet.max_row + 1):
                cell = sheet.cell(row=row, column=nodo_col)
                cell_value = str(cell.value or "").strip()
                
                # Si el valor coincide con el nodo buscado
                if cell_value == node_to_search or f"nodo {node_to_search}" == cell_value.lower():
                    results_found = True
                    
                    # Buscar placa y puerto en la misma fila
                    placa_value = None
                    port_value = None
                    
                    # Intentar detectar color de manera más exhaustiva
                    # 1. Buscar columnas que mencionen "color"
                    color = None
                    for col in range(1, sheet.max_column + 1):
                        header = str(sheet.cell(row=header_row, column=col).value or "").lower().strip()
                        if "color" in header:
                            color = self.get_cell_color(sheet, row, col)
                            break
                    
                    # 2. Si no se encontró, buscar en toda la fila por palabras clave
                    if not color:
                        for col in range(1, sheet.max_column + 1):
                            cell_value = str(sheet.cell(row=row, column=col).value or "").lower().strip()
                            if "verde" in cell_value:
                                color = "Verde"
                                break
                            elif "azul" in cell_value and not "tubo azul" in cell_value:
                                color = "Azul"
                                break
                            elif "naranja" in cell_value and not "tubo naranja" in cell_value:
                                color = "Naranja"
                                break
                    
                    # Revisar la misma fila para encontrar placa y puerto
                    for col in range(1, sheet.max_column + 1):
                        header = str(sheet.cell(row=header_row, column=col).value or "").lower().strip()
                        if "placa" in header:
                            placa_value = sheet.cell(row=row, column=col).value
                        elif "port" in header:
                            port_value = sheet.cell(row=row, column=col).value
                    
                    # Si no se encontró, buscar en columnas adyacentes
                    if not placa_value and not port_value:
                        for col in range(max(1, nodo_col - 5), min(nodo_col + 6, sheet.max_column + 1)):
                            col_value = str(sheet.cell(row=header_row, column=col).value or "").lower().strip()
                            if "placa" in col_value:
                                placa_value = sheet.cell(row=row, column=col).value
                            elif "port" in col_value or "puerto" in col_value:
                                port_value = sheet.cell(row=row, column=col).value
                    
                    # Determinar OLT con mejor detección de color
                    olt_info = self.get_olt_info(cell)
                    
                    # Mostrar resultados en formato compacto
                    self.results_text.insert(tk.END, "🎯 ", "success")
                    self.results_text.insert(tk.END, "Nodo encontrado: ", "success")
                    self.results_text.insert(tk.END, f"{node_to_search}\n", "info_value")
                    self.results_text.insert(tk.END, "─" * 60 + "\n", "separator")
                    
                    # Información en una sola línea para ahorrar espacio
                    self.results_text.insert(tk.END, "🔧 Placa: ", "label")
                    self.results_text.insert(tk.END, f"{placa_value}   ", "info_value")
                    self.results_text.insert(tk.END, "🔌 Puerto: ", "label")
                    self.results_text.insert(tk.END, f"{port_value}   ", "info_value") 
                    self.results_text.insert(tk.END, "🌐 OLT: ", "label")
                    self.results_text.insert(tk.END, f"{olt_info}   ", "info_highlight")
                    self.results_text.insert(tk.END, "📋 Hoja: ", "label")
                    self.results_text.insert(tk.END, f"{sheet_name}", "info_value")
        
        return results_found

    def analyze_sheet_structure(self, sheet):
        """Analiza la estructura de una hoja para identificar columnas importantes"""
        structure = {
            'nodo_columns': [],
            'placa_columns': {},
            'port_columns': {},
            'data_start_row': 2  # Asumimos que los datos comienzan en la fila 2 por defecto
        }
        
        # Buscar encabezados en las primeras 10 filas
        header_patterns = {
            'nodo': re.compile(r'nodo', re.I),
            'placa': re.compile(r'placa', re.I),
            'port': re.compile(r'(port|puerto)', re.I)
        }
        
        # Buscar en más filas para encontrar encabezados
        header_rows = []
        for row in range(1, min(15, sheet.max_row + 1)):
            header_candidates = 0
            for col in range(1, min(sheet.max_column + 1, 20)):  # Limitamos a las primeras 20 columnas por eficiencia
                cell_value = str(sheet.cell(row=row, column=col).value or "").strip()
                if any(pattern.search(cell_value) for pattern in header_patterns.values()):
                    header_candidates += 1
            
            if header_candidates > 0:
                header_rows.append(row)
        
        # Usar el último encabezado encontrado (más probable que sea el correcto)
        header_row = header_rows[-1] if header_rows else 1
        structure['data_start_row'] = header_row + 1
        
        # Identificar columnas importantes basadas en el encabezado
        for col in range(1, min(sheet.max_column + 1, 30)):  # Limitamos a 30 columnas para eficiencia
            cell_value = str(sheet.cell(row=header_row, column=col).value or "").strip().lower()
            
            if header_patterns['nodo'].search(cell_value):
                structure['nodo_columns'].append(col)
                
                # Buscar columnas de placa y puerto en un rango más amplio
                for offset in range(-5, 6):
                    check_col = col + offset
                    if 1 <= check_col <= sheet.max_column:
                        check_value = str(sheet.cell(row=header_row, column=check_col).value or "").strip().lower()
                        if header_patterns['placa'].search(check_value):
                            structure['placa_columns'][col] = check_col
                        elif header_patterns['port'].search(check_value):
                            structure['port_columns'][col] = check_col
        
        return structure

    def analizar_estructura_excel(self):
        """Analiza y muestra la estructura del Excel para diagnóstico"""
        if not self.excel_data:
            messagebox.showwarning("Advertencia", "Primero debes cargar un archivo Excel")
            return
            
        self.results_text.config(state="normal")  # Permitir edición temporalmente
        self.results_text.delete(1.0, tk.END)
        self.results_text.insert(tk.END, "=== ANÁLISIS DE ESTRUCTURA DEL EXCEL ===\n\n")
        
        for sheet_name in self.excel_data.sheetnames:
            sheet = self.excel_data[sheet_name]
            self.results_text.insert(tk.END, f"HOJA: {sheet_name}\n")
            self.results_text.insert(tk.END, f"- Filas: {sheet.max_row}\n")
            self.results_text.insert(tk.END, f"- Columnas: {sheet.max_column}\n\n")
            
            # Muestra las primeras 5 filas para ver la estructura
            self.results_text.insert(tk.END, "CONTENIDO DE LAS PRIMERAS 5 FILAS:\n")
            for row in range(1, min(6, sheet.max_row + 1)):
                row_data = []
                for col in range(1, min(10, sheet.max_column + 1)):  # Limitamos a 10 columnas
                    cell_value = str(sheet.cell(row=row, column=col).value or "")
                    row_data.append(cell_value[:20])  # Limitamos a 20 caracteres
                self.results_text.insert(tk.END, f"Fila {row}: {' | '.join(row_data)}\n")
            
            # Buscar columnas con "Nodo", "Placa" y "Port"
            self.results_text.insert(tk.END, "\nCOLUMNAS IMPORTANTES DETECTADAS:\n")
            for row in range(1, min(10, sheet.max_row + 1)):
                for col in range(1, min(30, sheet.max_column + 1)):
                    cell_value = str(sheet.cell(row=row, column=col).value or "").lower()
                    if "nodo" in cell_value:
                        self.results_text.insert(tk.END, f"- Columna NODO detectada: Col {col}, Fila {row}, Valor: {cell_value}\n")
                    elif "placa" in cell_value:
                        self.results_text.insert(tk.END, f"- Columna PLACA detectada: Col {col}, Fila {row}, Valor: {cell_value}\n")
                    elif "port" in cell_value:
                        self.results_text.insert(tk.END, f"- Columna PORT detectada: Col {col}, Fila {row}, Valor: {cell_value}\n")
            
            # Intentar detectar grupos de columnas (Nodo-Placa-Port)
            self.results_text.insert(tk.END, "\nGRUPOS DE COLUMNAS DETECTADOS:\n")
            header_patterns = {
                'nodo': re.compile(r'nodo', re.I),
                'placa': re.compile(r'placa', re.I),
                'port': re.compile(r'(port|puerto)', re.I)
            }
            for row in range(1, min(10, sheet.max_row + 1)):
                nodo_cols = []
                for col in range(1, min(30, sheet.max_column + 1)):
                    cell_value = str(sheet.cell(row=row, column=col).value or "")
                    if header_patterns['nodo'].search(cell_value):
                        nodo_cols.append(col)
                
                if nodo_cols:
                    self.results_text.insert(tk.END, f"Posible fila de encabezados: {row}\n")
                    for nodo_col in nodo_cols:
                        grupo = f"  Grupo: NODO en Col {nodo_col}"
                        for offset in range(1, 5):
                            if nodo_col + offset <= sheet.max_column:
                                check_col = nodo_col + offset
                                check_value = str(sheet.cell(row=row, column=check_col).value or "")
                                if header_patterns['placa'].search(check_value):
                                    grupo += f", PLACA en Col {check_col}"
                                elif header_patterns['port'].search(check_value):
                                    grupo += f", PORT en Col {check_col}"
                        self.results_text.insert(tk.END, f"{grupo}\n")
            
            self.results_text.insert(tk.END, "\n" + "-" * 50 + "\n\n")
        
        self.results_text.insert(tk.END, "Copia este texto y envíalo en tu siguiente mensaje para ayudarte mejor.\n")
        self.results_text.config(state="disabled")  # Deshabilitar edición nuevamente

if __name__ == "__main__":
    root = tk.Tk()
    app = NodoApp(root)
    root.mainloop()
    
#comando para compilar
#pyinstaller --onefile --windowed --icon=lupa.ico --name="Buscador_Nodos_FTTH" --add-data="lupa.ico;." --distpath="./dist" --workpath="./build" --specpath="." appnodos.py